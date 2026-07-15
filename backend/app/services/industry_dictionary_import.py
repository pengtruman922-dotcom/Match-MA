import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_ALIASES = {
    "l1": {"一级行业", "l1", "industry_l1"},
    "l2": {"二级行业", "l2", "industry_l2"},
    "aliases": {"别名", "映射词", "aliases"},
    "active": {"状态", "active"},
}


def parse_industry_import(file_name: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        rows = _csv_rows(content)
    elif suffix == ".xlsx":
        rows = _xlsx_rows(content)
    else:
        raise ValueError("Only .csv and .xlsx industry dictionary files are supported.")
    return _normalize_rows(rows)


def _csv_rows(content: bytes) -> list[list[Any]]:
    decoded = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ValueError("CSV encoding must be UTF-8 or GB18030.")
    return [list(row) for row in csv.reader(StringIO(decoded))]


def _xlsx_rows(content: bytes) -> list[list[Any]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _normalize_rows(rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not rows:
        raise ValueError("Import file is empty.")
    headers = [_text(value).lower() for value in rows[0]]
    indexes: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        index = next((position for position, header in enumerate(headers) if header in aliases), None)
        if index is not None:
            indexes[key] = index
    if "l1" not in indexes:
        raise ValueError("Import file must contain 一级行业 or l1 column.")

    normalized: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        l1 = _cell(row, indexes.get("l1"))
        l2 = _cell(row, indexes.get("l2"))
        aliases = _split_aliases(_cell(row, indexes.get("aliases")))
        active_text = _cell(row, indexes.get("active"))
        if not any((l1, l2, aliases, active_text)):
            continue
        normalized.append(
            {
                "row_number": row_number,
                "l1": l1,
                "l2": l2 or None,
                "aliases": aliases,
                "active": active_text.lower() not in {"停用", "禁用", "false", "0", "no"},
            }
        )
    if not normalized:
        raise ValueError("Import file contains no data rows.")
    return normalized


def _cell(row: list[Any], index: int | None) -> str:
    return _text(row[index]) if index is not None and index < len(row) else ""


def _text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _split_aliases(value: str) -> list[str]:
    items: list[str] = []
    normalized = value.replace("，", ",").replace("、", ",").replace(";", ",").replace("；", ",")
    for raw in normalized.split(","):
        item = raw.strip()
        if item and item not in items:
            items.append(item)
    return items
